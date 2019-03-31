#----------------------------------------------------------------------------------
# TRD xlswriter class
# Parent Class for xls writer of training data
#----------------------------------------------------------------------------------

#-------------------------------------------------------------------------
# 
# (C) 2019, jdmorise at yahoo.com
#  
# 	 This program is free software: you can redistribute it and/or modify
# 	 it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <https://www.gnu.org/licenses/>.
#
#
#-------------------------------------------------------------------------


from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.drawing.text import CharacterProperties


class trd_xlswriter: 

    def __init__(self,xlsx_filename):
        
        self.xlsx_filename = xlsx_filename
        
        # Header and overview
        self.header_font = Font(size=12, bold=True) 
        self.header_align = Alignment(horizontal='center')
        # Data
        
        self.data_font = Font(size=12, bold=False) 
        self.data_align = Alignment(horizontal='center', vertical='center')

        # Border Type
        self.thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
                
    # Subfunctions
    def format_header(self, min_col, min_row, max_col, max_row): 
        for row in self.ws.iter_rows(min_col=min_col,  min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.font = self.header_font
                cell.border = self.thin_border  
                cell.alignment = self.header_align
                                
    def format_data(self, min_col, min_row, max_col, max_row): 
        for row in self.ws.iter_rows(min_col=min_col,  min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.font = self.data_font
                cell.border = self.thin_border  
                cell.alignment = self.data_align
                
    #+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    ## Save XLSX File
    
    def save(self): 
        self.wb.save(self.xlsx_filename)
        



        
        